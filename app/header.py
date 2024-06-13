import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

def create_logo(root, app):
    logo_frame = tk.Frame(root, bg='#333333')
    logo_frame.pack(side=tk.TOP, pady=10, padx=35, fill=tk.X)
    
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    logo_path = os.path.join(BASE_DIR, 'assets', 'logo.png')
    
    if not os.path.exists(logo_path):
        print(BASE_DIR + 'cosa')
        raise FileNotFoundError(f"No se encontró el archivo: {logo_path}")

    logo_image = tk.PhotoImage(file=logo_path)
    logo_label = tk.Label(logo_frame, image=logo_image, bg='#333333')
    logo_label.image = logo_image
    logo_label.pack(side=tk.LEFT)

    title_label = tk.Label(logo_frame, text="Gestor de Permisos", font=("Helvetica", 16, "bold"), bg='#333333', fg='white')
    title_label.pack(side=tk.LEFT, padx=10)

    tk.Label(logo_frame, text="", bg='#333333').pack(side=tk.LEFT, expand=True)

    redirect_button = tk.Button(logo_frame, text="MANUAL", command=lambda: app.redirect_to_new_page(), bg="lightgrey", fg="black", font=("Helvetica", 10, "bold"))
    redirect_button.pack(side=tk.LEFT, padx=10)

    download_button = tk.Button(logo_frame, text="DESCARGAR PLANTILLA", command=download_excel, bg="#00E51F", fg="black", font=("Helvetica", 10, "bold"))
    download_button.pack(side=tk.LEFT, padx=10)

    return logo_frame

def download_excel():
    data = {
        'Directorios \ Grupos': [
            '/home/ejemplo/Desktop/Proyecto1', 
            '/home/ejemplo/Desktop/Proyecto2', 
            '/home/ejemplo/Desktop/Proyecto3'
        ],
        'Otros': ['---', '---', '---'],
        'Grupo1': ['rxw', 'rxw', 'rxw'],
        'Grupo2': ['r-x', 'r-x', 'r-x'],
        'Grupo3': ['---', '---', '---'],
        'Grupo4': ['---', '---', '---'],
    }
    df = pd.DataFrame(data)

    # Abrir un diálogo para guardar el archivo
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if file_path:
        df.to_excel(file_path, index=False)
        
        # Cargar el archivo para aplicar estilos
        wb = load_workbook(file_path)
        ws = wb.active

        # Aplicar estilos a los encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

        # Ajustar ancho de las columnas
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(file_path)
        print(f"Archivo guardado en {file_path}")

import tkinter as tk
import ctypes

class NewPage:
    def __init__(self, root):
        self.root = root
        self.root.title("Manual de Permisos")
        self.set_fixed_window_size()
        self.root.configure(bg='#333333')

        # Crear un contenedor con un canvas y una scrollbar
        self.canvas = tk.Canvas(self.root, bg='#333333', highlightthickness=0)
        self.scrollbar = tk.Scrollbar(self.root, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg='#333333')

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        self.create_widgets()
    
    def set_fixed_window_size(self):
        fixed_width = 700 
        fixed_height = 600 
        self.root.geometry(f"{fixed_width}x{fixed_height}")

    def create_widgets(self):
        tk.Label(self.scrollable_frame, text="Manual ACL", font=("Helvetica", 16, "bold"), bg='#333333', fg='white').pack(pady=10)

        buttons_info = [
            ("CONSULTAR", "Consulta los permisos actuales del directorio. Estos deberían corresponder con los mostrados en el excel. En caso de que no correspondan, se deben setear según la opción que considere más adecuada."),
            ("SETEAR", "Devuelve un comando para setear cada uno de los permisos al directorio seleccionado."),
            ("SETEAR R", "Devuelve un comando para setear cada uno de los permisos al directorio seleccionado de manera recursiva (Continuando con el mismo proceso con los subdirectorios y archivos internos)."),
            ("ELIMINAR", "Devuelve un comando para eliminar los permisos del directorio seleccionado (Incluido los permisos por defecto)"),
            ("ELIMINAR R", "Devuelve un comando para eliminar los permisos del directorio seleccionado de manera recursiva (Continuando con el mismo proceso con los subdirectorios y archivos internos)"),
            ("SETEAR POR DEFECTO", "Devuelve un comando para setear los permisos por defecto (Los que se agregan en ese directorio)"),
            ("SETEAR POR DEFECTO R", "Devuelve un comando para setear los permisos por defecto de manera recursiva (Esto se aplica a todos los directorios internas) ")
        ]

        for btn_text, description in buttons_info:
            frame = tk.Frame(self.scrollable_frame, bg='#333333')
            frame.pack(pady=20, padx=10, fill=tk.X)

            button = tk.Button(frame, text=btn_text, bg="#444444", fg="white", font=("Helvetica", 10, "bold"), width=25, relief="flat")
            button.grid(row=0, column=0, padx=5, pady=0, sticky="n")

            label = tk.Label(frame, text=description, font=("Helvetica", 10), bg='#333333', fg='white', wraplength=400, justify="left")
            label.grid(row=0, column=1, padx=5, pady=0, sticky="w")

            frame.grid_columnconfigure(1, weight=1)


if __name__ == "__main__":
    root = tk.Tk()
    app = NewPage(root)
    root.mainloop()
